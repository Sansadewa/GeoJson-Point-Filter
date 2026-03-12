#!/usr/bin/env python3
"""
Import rekap1_081329.xlsx → Supabase sbr_data table.

Reads each sheet in chunks using openpyxl to avoid loading 635K rows
into memory all at once. Batches 500 rows per Supabase API call.

All 3 sheets imported with a 'sheet_name' column to track origin.

Usage:
    python scripts/import_sbr_data.py rekap1_081329.xlsx
    python scripts/import_sbr_data.py rekap1_081329.xlsx --sheet "Sudah GC"
    python scripts/import_sbr_data.py rekap1_081329.xlsx --dry-run
    python scripts/import_sbr_data.py rekap1_081329.xlsx --resume 1500
"""

import sys
import re
import json
import time
import math
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from supabase import create_client

BATCH_SIZE  = 500
READ_CHUNK  = 5000   # rows to read from Excel at a time
SHEETS      = ["Sudah GC", "Belum GC", "Duplikat"]

EXPECTED_COLS = [
    "idsbr", "nama_usaha", "alamat_usaha", "kegiatan_usaha", "skala_usaha",
    "sumber_data", "kode_wilayah", "kdprov", "kdkab", "kdkec", "kddesa",
    "nmprov", "nmkab", "nmkec", "nmdesa", "latitude", "longitude",
    "latlong_status", "gcs_result", "status_gc", "status_perusahaan",
    "gc_username", "latitude_gc", "longitude_gc", "latlong_status_gc",
    "history_ref_profiling_id", "perusahaan_id", "captured_at", "batch_start",
]


def load_secrets():
    secrets_path = Path(__file__).parent.parent / ".streamlit" / "secrets.toml"
    text = secrets_path.read_text()
    url = re.search(r'url\s*=\s*"([^"]+)"', text).group(1)
    key = re.search(r'key\s*=\s*"([^"]+)"', text).group(1)
    return url, key


def clean_val(v, col: str = "") -> str | None:
    """Normalise a single cell value to str or None.
    All values stored as TEXT; the app's clean_coord() handles coordinate parsing.
    """
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "nan", "None", "NaT", "NaN", "<NA>"):
        return None
    return s


def iter_sheet_chunks(excel_path: str, sheet_name: str):
    """
    Yield (col_names, list_of_row_dicts) for each READ_CHUNK rows.
    Uses openpyxl read_only mode — never loads the whole sheet.
    The header row (containing column names) is expected at Excel row 2 (1-indexed).
    Data starts at Excel row 3.
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    col_names = None
    chunk = []
    data_row_count = 0

    for excel_row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if excel_row_num == 1:
            # Row 1 is blank / unnamed — skip
            continue
        if excel_row_num == 2:
            # Row 2 is the header
            col_names = [str(c).strip() if c is not None else f"col_{i}"
                         for i, c in enumerate(row)]
            continue

        # Row 3+ is data
        data_row_count += 1
        record = {col_names[i]: clean_val(v, col_names[i]) for i, v in enumerate(row) if i < len(col_names)}

        # Skip rows where idsbr is empty (phantom rows)
        if not record.get("idsbr"):
            continue

        chunk.append(record)

        if len(chunk) >= READ_CHUNK:
            yield col_names, chunk
            chunk = []

    if chunk:
        yield col_names, chunk

    wb.close()


def import_sheet(client, excel_path: str, sheet_name: str,
                 dry_run: bool, resume_row: int) -> int:
    """Stream one sheet and insert rows in batches. Returns rows inserted."""

    print(f"\n{'='*60}")
    print(f"Sheet: '{sheet_name}'")
    print(f"{'='*60}", flush=True)

    inserted   = 0
    errors     = 0
    batch_buf  = []   # accumulate rows across chunks before sending
    global_row = 0    # count of data rows seen
    batch_num  = 0
    t_start    = time.time()

    for col_names, chunk in iter_sheet_chunks(excel_path, sheet_name):
        # Validate columns on first chunk
        if inserted == 0 and errors == 0 and not batch_buf:
            missing = [c for c in EXPECTED_COLS if c not in col_names]
            if missing:
                print(f"ERROR: Missing columns: {missing}")
                print(f"Found: {col_names}")
                return 0
            print(f"Columns : ✅ all {len(EXPECTED_COLS)} confirmed")
            if dry_run:
                print(f"DRY RUN : sample row = {dict(list(chunk[0].items())[:4])}...")

        for record in chunk:
            global_row += 1

            # Skip rows before resume point
            if global_row <= resume_row:
                continue

                # Keep only the columns that exist in the table
                record = {k: record.get(k) for k in EXPECTED_COLS}
            batch_buf.append(record)

            if len(batch_buf) >= BATCH_SIZE:
                batch_num += 1
                if not dry_run:
                    ok = _send_batch(client, batch_buf, batch_num)
                    if ok:
                        inserted += len(batch_buf)
                    else:
                        errors += 1
                else:
                    inserted += len(batch_buf)
                batch_buf = []

                # Progress
                elapsed = time.time() - t_start
                rate    = inserted / elapsed if elapsed > 0 else 1
                print(
                    f"\r  batch {batch_num:>5}  {inserted:>8,} rows  "
                    f"{rate:6.0f} rows/s  elapsed {elapsed/60:4.1f}m",
                    end="", flush=True
                )

    # Flush remaining rows
    if batch_buf:
        batch_num += 1
        if not dry_run:
            ok = _send_batch(client, batch_buf, batch_num)
            if ok:
                inserted += len(batch_buf)
            else:
                errors += 1
        else:
            inserted += len(batch_buf)
        batch_buf = []

    elapsed = time.time() - t_start
    action  = "Would insert" if dry_run else "Inserted"
    print(f"\n  {action} {inserted:,} rows in {elapsed/60:.1f} min  ({errors} error batches)")
    return inserted


def _send_batch(client, rows: list, batch_num: int) -> bool:
    """Send one batch with retry. Returns True on success, False on failure."""
    for attempt in range(3):
        try:
            client.table("sbr_data").insert(rows).execute()
            return True
        except Exception as e:
            if attempt < 2:
                time.sleep(2 ** attempt)
            else:
                print(f"\n  ERROR batch {batch_num}: {e}")
                log = Path(f"import_errors_batch{batch_num}.jsonl")
                with open(log, "a") as f:
                    f.write(json.dumps({"batch": batch_num, "error": str(e),
                                        "first_idsbr": rows[0].get("idsbr")}) + "\n")
                return False
    return False


def main():
    parser = argparse.ArgumentParser(description="Import rekap Excel → Supabase sbr_data")
    parser.add_argument("excel_file")
    parser.add_argument("--sheet",   default=None, help="Import only this sheet")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--resume",  type=int, default=0, metavar="ROW",
                        help="Skip first N data rows (single-sheet mode only)")
    args = parser.parse_args()

    sheets = [args.sheet] if args.sheet else SHEETS

    print(f"{'='*60}")
    print(f"SBR Data Importer")
    print(f"{'='*60}")
    print(f"File    : {args.excel_file}")
    print(f"Sheets  : {sheets}")
    print(f"Dry run : {args.dry_run}")

    url, key   = load_secrets()
    client     = create_client(url, key)

    try:
        client.table("sbr_data").select("idsbr").limit(1).execute()
        print("Supabase : ✅ connected")
    except Exception as e:
        print(f"Supabase : ❌ {e}")
        sys.exit(1)

    grand_total = 0
    t_all       = time.time()

    for sheet in sheets:
        resume = args.resume if (args.sheet and len(sheets) == 1) else 0
        count  = import_sheet(client, args.excel_file, sheet,
                               dry_run=args.dry_run, resume_row=resume)
        grand_total += count

    elapsed = time.time() - t_all
    print(f"\n{'='*60}")
    action = "Would insert" if args.dry_run else "Total inserted"
    print(f"{action}: {grand_total:,} rows in {elapsed/60:.1f} min")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
